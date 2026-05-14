import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import tempfile
import time

st.set_page_config(layout="wide")

st.title("Excel Processor")

# =========================
# FILE UPLOAD SECTION
# =========================

st.subheader("Upload Files")

uploaded_file = st.file_uploader(
    "Upload Excel File",
    type=["xlsx"]
)

st.caption(
    """
Required Sheets:
- Bangalore
- Chennai
- Hyderabad
- Mumbai
- Kolkata
- NCR

Required Columns:
HUB, Location, Zone/COC, Owner, Customer Code, Customer Name,
Order No, Invoice No, WF_TaskID, Shap Hrs., Performed Hrs,
Billed Hrs, Variance, Excess Paid, Excess billing, Short billing,
Short / Missing Roster, Training & OJT, Complimentary Hrs., BFL Remarks

Note:
- Headers can be located anywhere within the first 10 rows.
- Tool automatically searches first 10 rows to detect headers.
"""
)

# =========================
# RUN BUTTON
# =========================

run_button = st.button("Run")

# =========================
# PROCESSING
# =========================

if run_button:

    log_container = st.container()

    with log_container:

        status_text = st.empty()
        progress_bar = st.progress(0)

        if uploaded_file is None:
            st.error("Please upload an Excel file.")
            st.stop()

        # =========================
        # SAVE TEMP FILE
        # =========================

        status_text.info("Initializing file processing...")
        progress_bar.progress(5)
        time.sleep(0.2)

        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(uploaded_file.read())
                file_path = tmp.name
        except Exception as e:
            st.error(f"Error while saving uploaded file: {e}")
            st.stop()

        # =========================
        # REQUIRED CONFIG
        # =========================

        required_sheets = [
            'Bangalore',
            'Chennai',
            'Hyderabad',
            'Mumbai',
            'Kolkata',
            'NCR'
        ]

        desired_columns = [
            'HUB',
            'Location',
            'Zone/COC',
            'Owner',
            'Customer Code',
            'Customer Name',
            'Order No',
            'Invoice No',
            'WF_TaskID',
            'Shap Hrs.',
            'Performed Hrs',
            'Billed Hrs',
            'Variance',
            'Excess Paid',
            'Excess billing',
            'Short billing',
            'Short / Missing Roster',
            'Training & OJT',
            'Complimentary Hrs.',
            'BFL Remarks'
        ]

        desired_columns_lower = [col.lower() for col in desired_columns]

        cols_to_fix_master = [
            'Excess Paid',
            'Excess Billing',
            'Short Billing',
            'Short / Missing Roster',
            'Training & Ojt',
            'Complimentary Hrs.'
        ]

        # =========================
        # HEADER FINDER FUNCTION
        # =========================

        def read_excel_with_dynamic_header(file_path, sheet_name):

            try:
                preview_df = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    header=None,
                    nrows=10
                )
            except Exception as e:
                raise Exception(f"Unable to preview sheet '{sheet_name}': {e}")

            header_row = None

            for i in range(min(10, len(preview_df))):

                row_values = (
                    preview_df.iloc[i]
                    .astype(str)
                    .str.strip()
                    .str.lower()
                    .tolist()
                )

                matched_cols = [
                    col for col in desired_columns_lower
                    if col in row_values
                ]

                if len(matched_cols) >= 5:
                    header_row = i
                    break

            if header_row is None:
                raise Exception(
                    f"Header row not found within first 10 rows in sheet '{sheet_name}'."
                )

            try:
                df = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    header=header_row
                )
            except Exception as e:
                raise Exception(f"Unable to read sheet '{sheet_name}': {e}")

            return df

        # =========================
        # VALIDATE SHEETS
        # =========================

        status_text.info("Validating workbook sheets...")
        progress_bar.progress(10)
        time.sleep(0.2)

        try:
            excel_file = pd.ExcelFile(file_path)
            existing_sheets = excel_file.sheet_names
        except Exception as e:
            st.error(f"Error reading workbook: {e}")
            st.stop()

        missing_sheets = [
            sheet for sheet in required_sheets
            if sheet not in existing_sheets
        ]

        if missing_sheets:
            st.error(f"Required sheet(s) not found: {missing_sheets}")
            st.stop()

        # =========================
        # READ SHEETS
        # =========================

        dataframes = {}

        for idx, sheet_name in enumerate(required_sheets):

            status_text.info(
                f"Reading and detecting headers in sheet: {sheet_name}"
            )

            progress_value = 15 + int((idx / len(required_sheets)) * 25)
            progress_bar.progress(progress_value)

            time.sleep(0.2)

            try:
                df = read_excel_with_dynamic_header(file_path, sheet_name)
                dataframes[sheet_name] = df
            except Exception as e:
                st.error(f"Error processing sheet '{sheet_name}': {e}")
                st.stop()

        # =========================
        # PROCESS SHEETS
        # =========================

        status_text.info("Processing sheet data...")
        progress_bar.progress(45)
        time.sleep(0.2)

        for name, df in dataframes.items():

            try:
                df.columns = df.columns.str.strip().str.lower()
            except Exception as e:
                st.error(f"Error cleaning column names in sheet '{name}': {e}")
                st.stop()

            # Validate required columns
            missing_cols = [
                col for col in desired_columns_lower
                if col not in df.columns
            ]

            if missing_cols:
                st.error(
                    f"Missing columns in sheet '{name}': {missing_cols}"
                )
                st.stop()

            try:
                existing_cols = [
                    col for col in desired_columns_lower
                    if col in df.columns
                ]

                df = df[existing_cols]
            except Exception as e:
                st.error(
                    f"Error selecting required columns in sheet '{name}': {e}"
                )
                st.stop()

            try:
                df.columns = df.columns.str.title()
                df = df.copy()
            except Exception as e:
                st.error(
                    f"Error formatting columns in sheet '{name}': {e}"
                )
                st.stop()

            try:
                df['Customer Code'] = (
                    df['Customer Code']
                    .fillna('')
                    .astype(str)
                    .str.replace(r'\.0$', '', regex=True)
                    .str.strip()
                )
            except Exception as e:
                st.error(
                    f"Error cleaning Customer Code in sheet '{name}': {e}"
                )
                st.stop()

            try:
                df = df[df['Customer Code'] != '']
            except Exception as e:
                st.error(
                    f"Error filtering blank Customer Code rows in sheet '{name}': {e}"
                )
                st.stop()

            try:
                df['Loc'] = name
            except Exception as e:
                st.error(
                    f"Error creating Loc column in sheet '{name}': {e}"
                )
                st.stop()

            try:
                cols = ['Loc'] + [col for col in df.columns if col != 'Loc']
                df = df[cols]
            except Exception as e:
                st.error(
                    f"Error reordering columns in sheet '{name}': {e}"
                )
                st.stop()

            try:
                cols_to_fix = [
                    col for col in cols_to_fix_master
                    if col in df.columns
                ]
            except Exception as e:
                st.error(
                    f"Error identifying cleanup columns in sheet '{name}': {e}"
                )
                st.stop()

            # Clean properly
            for col in cols_to_fix:

                try:
                    df[col] = (
                        df[col]
                        .replace([0, 0.0], pd.NA)
                        .astype(str)
                        .replace(['nan', 'None'], '')
                        .str.strip()
                    )
                except Exception as e:
                    st.error(
                        f"Error cleaning column '{col}' in sheet '{name}': {e}"
                    )
                    st.stop()

            try:
                df[cols_to_fix] = df[cols_to_fix].replace('', pd.NA)
            except Exception as e:
                st.error(
                    f"Error converting blanks to NA in sheet '{name}': {e}"
                )
                st.stop()

            try:
                df = df.dropna(subset=cols_to_fix, how='all')
            except Exception as e:
                st.error(
                    f"Error filtering rows in sheet '{name}': {e}"
                )
                st.stop()

            dataframes[name] = df

        # =========================
        # COMBINE DATA
        # =========================

        status_text.info("Combining all city data...")
        progress_bar.progress(70)
        time.sleep(0.2)

        try:
            all_cities_df = pd.concat(
                dataframes.values(),
                ignore_index=True
            )
        except Exception as e:
            st.error(f"Error combining sheets: {e}")
            st.stop()

        try:
            all_cities_df['Loc'] = all_cities_df['Loc'].mask(
                all_cities_df.duplicated(subset=['Loc']),
                ''
            )
        except Exception as e:
            st.error(f"Error formatting Loc column: {e}")
            st.stop()

        for col in cols_to_fix_master:

            if col in all_cities_df.columns:

                try:
                    all_cities_df[col] = pd.to_numeric(
                        all_cities_df[col],
                        errors='coerce'
                    )
                except Exception as e:
                    st.error(
                        f"Error converting column '{col}' to numeric: {e}"
                    )
                    st.stop()

        # =========================
        # WRITE OUTPUT
        # =========================

        status_text.info("Generating output workbook...")
        progress_bar.progress(85)
        time.sleep(0.2)

        try:
            output_path = tempfile.NamedTemporaryFile(
                delete=False,
                suffix=".xlsx"
            ).name
        except Exception as e:
            st.error(f"Error creating output file: {e}")
            st.stop()

        try:
            wb = load_workbook(file_path)
        except Exception as e:
            st.error(f"Error loading workbook for output: {e}")
            st.stop()

        try:
            if 'All Cities' in wb.sheetnames:
                del wb['All Cities']

            ws = wb.create_sheet(title='All Cities')
        except Exception as e:
            st.error(f"Error creating output sheet: {e}")
            st.stop()

        try:
            for c_idx, col_name in enumerate(all_cities_df.columns, start=1):
                ws.cell(row=1, column=c_idx, value=col_name)
        except Exception as e:
            st.error(f"Error writing headers to output: {e}")
            st.stop()

        try:
            for r_idx, row in enumerate(
                all_cities_df.itertuples(index=False),
                start=2
            ):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
        except Exception as e:
            st.error(f"Error writing data to output workbook: {e}")
            st.stop()

        try:
            wb.save(output_path)
        except Exception as e:
            st.error(f"Error saving output workbook: {e}")
            st.stop()

        # =========================
        # SUCCESS
        # =========================

        status_text.success("Processing completed successfully.")
        progress_bar.progress(100)

        try:
            with open(output_path, "rb") as f:

                st.download_button(
                    label="Download Output",
                    data=f,
                    file_name="final_output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        except Exception as e:
            st.error(f"Error preparing download file: {e}")
            st.stop()

# =========================
# DOCUMENTATION SECTION
# =========================

with st.expander("What This Tool Does"):

    st.write(
        """
This tool processes multiple city-wise Excel sheets and combines them into
a single consolidated report.

The application:
- Reads data from all required city sheets
- Automatically detects headers within first 10 rows
- Cleans and standardizes data
- Filters only meaningful variance-related rows
- Combines all city data into a unified output sheet
- Generates a downloadable Excel report
"""
    )

with st.expander("How to Use"):

    st.write(
        """
1. Upload the Excel file  
2. Click the Run button  
3. Wait for processing to complete  
4. Download the final consolidated report  
"""
    )

with st.expander("Output Details"):

    st.write(
        """
The output workbook contains:
- Original sheets from uploaded workbook
- Additional consolidated sheet named:
  "All Cities"

The report structure includes:
HUB → Location → Zone/COC → Owner → Customer → Order → Invoice

Only rows containing variance-related information are retained in final output.
"""
    )

with st.expander("Financial Logic"):

    st.write(
        """
The tool identifies and consolidates operational and billing exceptions.

Key checks include:
- Excess Paid
- Excess Billing
- Short Billing
- Missing Roster Issues
- Training & OJT Hours
- Complimentary Hours

Rows without any financial or operational variance indicators are removed
to keep the report focused and actionable.
"""
    )
